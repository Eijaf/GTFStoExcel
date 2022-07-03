# -*- coding: utf-8 -*-
"""
Created on Tue Jun 21 21:34:30 2022

@author: Adrien SCHWEGLER
"""
import csv, sqlite3,os
from openpyxl import load_workbook, Workbook
from copy import deepcopy


def extratDataFromFile(fileName, columnName): 
     with open(fileName, 'r', encoding='utf-8_sig') as file:
            dr = csv.DictReader(file)
            table_info = [tuple([i[column] for column in columnName]) for i in dr]
     return table_info
 
    
def createDataBaseTableFromCsv(csvfile):
    indexStops, indexDirection = -1, -1  #réinitialisation de l'index
        
    #Récupèration des colones du fichier
    with open(csvfile, 'r', encoding='utf-8_sig') as file:
        listColumn = file.readline().rstrip()      #first line of the file
        columnName = listColumn.split(",")          #from str to list

    
    table_info = extratDataFromFile(csvfile, columnName)  #contenue de la table
    print('        -Information de la table récupérée')

    
    numberOfColumn = len(columnName)  #nombre de colone    
    createSintax = listColumn   #str for insertion in CREATE sentence
    #add INTERGER when needed
    indexStops = createSintax.find('stop_sequence')
    indexDirection = createSintax.find('direction_id')
    if indexStops != -1:
        createSintax = createSintax[:indexStops+13] + ' INTERGER ' + createSintax[indexStops+13:]
    if indexDirection != -1:
        createSintax = createSintax[:indexDirection+12] + ' INTERGER ' + createSintax[indexDirection+12:]
    #create and fill table
    cursor = connection.cursor()
    cursor.execute("CREATE TABLE {} ({});".format(csvfile[:-4], createSintax))
    cursor.executemany("INSERT INTO {} ({}) VALUES ({});".format(csvfile[:-4], listColumn, ('?,'*numberOfColumn)[:-1]), table_info)  #Outch
    connection.commit()

def extractTrip(listRoutes, direction):
    dictTrip = {}
    for route in listRoutes:
        cursor = connection.cursor()
        cursor.execute('''SELECT trips.trip_id, service_id
                       FROM trips
                       JOIN stop_times
                       ON stop_times.trip_id = trips.trip_id
                       WHERE trips.route_id = "{}" AND trips.direction_id = "{}"
                       ORDER BY stop_times.departure_time
                        '''.format(route, direction))
        dictTrip[route] = {key : {} for key in cursor}  #je prend à la fois trip_id et service id.
    return dictTrip


def GetstrDays(service_id):
    cursor = connection.cursor()
    cursor.execute('''SELECT monday, tuesday, wednesday, thursday, friday, saturday, sunday
                   FROM calendar
                   WHERE service_id = "{}"'''.format(service_id))
    listDays = cursor.fetchall()
    strday = ''
    for day in listDays: #on écrit L, M, W... en fonction de si le bus passe tel jours
        if day[0] == '1':
            strday += 'L'
        if day[1] == '1':
            strday += 'Ma'
        if day[2] == '1':
            strday += 'Me'
        if day[3] == '1':
            strday += 'J'
        if day[4] == '1':
            strday += 'V'
        if day[5] == '1':
            strday += 'S'
        if day[6] == '1':
            strday += 'D'
    return strday

def rewriteStrDAys(dictPeriodeTrips):
    for trips, strday in dictPeriodeTrips.items():
        if strday == 'LMaMeJV':
            dictPeriodeTrips[trips] = 'LàV'
        elif strday == 'LMaMeJVS':
            dictPeriodeTrips[trips] = 'LàS'
        elif strday == 'Me':
            dictPeriodeTrips[trips] = 'Mer'
    return dictPeriodeTrips      

def GetstrPeriode(service_id) :
    cursor = connection.cursor()
    cursor.execute('''SELECT start_date, end_date
                   FROM calendar
                   WHERE service_id = "{}"'''.format(service_id))
    periodBrut =  cursor.fetchall()[0]
    start, end = periodBrut[0], periodBrut[1]
    return (start[-2:] + '/'+ start[4:6]+ '/' + start[:4], end[-2:] + '/'+end[4:6]+ '/' + end[:4])   #change le format de la date          

def findTripNumber(trip_id):
    index_end_number = 0   
    for caract in trip_id:      #looking for the len of the trips's number at the begining of the trip_id
        if 48<=ord(caract)<=57: #if it is a number
            index_end_number += 1 #add one to the index of the end
        else:
            break#if not : the number has end so break the loop
    return trip_id[:index_end_number]       

def suppdoubleAndExtractStops(dictTrip, listRoute): #from trips_id And extrtac days
    dictTripStops = deepcopy(dictTrip)
    dicDays = {} #servira a stoké les trips de route avec les jour de passage
    dicPeriode = {}
    for route_id, value in dictTrip.items():
        compare = []                #list of all trips number already pass(not necessarly same day)
        dictDaysTrips = {}
        dictPeriodeTrips = {}
        for tripAndService_id, noth in value.items():
            trip_id = tripAndService_id[0]
            service_id =  tripAndService_id[1]
            tripNumber = findTripNumber(trip_id)                     
            if tripNumber in compare:                
                #ajouter jours que j'enleve au même num dans le dictionnaire a numéro comme clé
                
                
                dictDaysTrips[tripNumber] += GetstrDays(service_id) #+= cause already exist
                
                
                del dictTripStops[route_id][tripAndService_id] #delet the double
            else :
               cursor = connection.cursor()
               cursor.execute('''SELECT stops.stop_id, stop_times.departure_time
                              FROM stops
                              JOIN stop_times
                              ON stop_times.stop_id = stops.stop_id 
                              WHERE stop_times.trip_id = "{}" AND stop_times.drop_off_type !="1"
                              ORDER BY stop_times.stop_sequence'''.format(trip_id))
               #ET récupérer jours dans un dictionnaire avec pour clé le numéro devant le trip id et ROUTE ID pour pouvoir l'utiliser au bon moment
               for stops in cursor.fetchall():
                   dictTripStops[route_id][tripAndService_id][stops[0]] = stops[1][:5]
               dictDaysTrips[tripNumber] = GetstrDays(service_id)
               dictPeriodeTrips[tripNumber] = GetstrPeriode(service_id)
               compare.append(tripNumber)
         
        dictDaysTrips = rewriteStrDAys(dictDaysTrips) #on réécrit le cat régulier des jours
        dicDays[route_id] = dictDaysTrips
        dicPeriode[route_id] = dictPeriodeTrips
    return dictTripStops, dicDays, dicPeriode


def extractStops(route, direction): #from route_id
    cursor = connection.cursor()
    cursor. execute('''SELECT stops.stop_id ,stops.stop_name 
                    FROM stops
                    JOIN stop_times
                    ON stop_times.stop_id = stops.stop_id
                    JOIN trips
                    ON trips.trip_id = stop_times.trip_id
                    WHERE trips.route_id = "{}" AND trips.direction_id = {} AND stop_times.drop_off_type !="1"
                    GROUP BY stops.stop_id
                    ORDER BY stop_times.stop_sequence ASC'''.format(route, direction))
    return {row[0] : row[1] for row in cursor} 


def createTable(route, dictStops, dictSens, dicDayTrip, dicPeriodTrip) :
    table = [['Début de validité'], ['Fin de validité'], ['Jours de passage']] + [[stops] for stops in dictStops]
    dictRoute = dictSens[route]    
    for tripAndService, stops in dictRoute.items():
        trip_id = tripAndService[0]
        tripNumber = findTripNumber(trip_id)        
        table[0].append(dicPeriodTrip[tripNumber][0]) #début de validité
        table[1].append(dicPeriodTrip[tripNumber][1]) #fin de validité
        table[2].append(dicDayTrip[tripNumber]) #jours de passage du voyage        
        
        for i in range(3, len(dictStops)+3):     #heure de passage du voyage
            if table[i][0] in stops: 
                table[i].append(stops[table[i][0]])
            else :
                table[i].append('-')
    for stops in table: #je change le stop_id pour le stop_name
        try :
            stops[0] = dictStops[stops[0]]
        except:  #to avoid key problem when at 'jour de validité'
            pass
    return table            
    
def createCSV(dictSens0, dictSens1, dicDays0, dicDays1, dicPeriodRoute0, dicPeriodRoute1):
    for route in listRoutes:
        fileName = route + '.csv' 
        dictStops0 = extractStops(route, 0) 
        dictStops1 = extractStops(route, 1)
        dicDayTrip0, dicDayTrip1 = dicDays0[route], dicDays1[route]
        dicPeriodTrip0, dicPeriodTrip1 = dicPeriodRoute0[route], dicPeriodRoute1[route]
        table0 = createTable(route, dictStops0, dictSens0, dicDayTrip0, dicPeriodTrip0) #grille sens 0
        table1 = createTable(route, dictStops1, dictSens1, dicDayTrip1, dicPeriodTrip1) #grille sens 1
        with open(fileName,'w', newline = '', encoding='utf-8') as file:
            obj = csv.writer(file, delimiter = ",")
            for element in table0:  #on écrit le sens 0
                obj.writerow(element)
            obj.writerow('\n')
            for element in table1:  #on écrit le sens 1
                obj.writerow(element)


def createXLS(listRoutes, dictSens0, dictSens1,dicDays0, dicDays1, dicPeriodRoute0, dicPeriodRoute1):
    wb = Workbook()
    wb.save('Horraires_GTFS.xlsx')
    wb = load_workbook('Horraires_GTFS.xlsx')
    sheet = wb.active
    sheet.title = 'Informations'
    sheet['A1'] = 'Ce fichier contient les routes :'
    
    for element in listRoutes:
         sheet.append([element])
    compteur = 1
    nb_routes = len(listRoutes)
    for route in listRoutes : #création des feuilles
        print('    -Feuille {} sur {}.'.format(compteur, nb_routes))
        dictStops0 = extractStops(route, 0) 
        dictStops1 = extractStops(route, 1)
        dicDayTrip0, dicDayTrip1 = dicDays0[route], dicDays1[route]
        dicPeriodTrip0, dicPeriodTrip1 = dicPeriodRoute0[route], dicPeriodRoute1[route]
        table0 = createTable(route, dictStops0, dictSens0, dicDayTrip0, dicPeriodTrip0) #grille sens 0
        table1 = createTable(route, dictStops1, dictSens1, dicDayTrip1, dicPeriodTrip1) #grille sens 1
        wb.create_sheet(title=route)
        sheet = wb[route]
        for i in table0:
            sheet.append(i) 
        sheet.append([]) 
        for i in table1:
            sheet.append(i) 
        compteur +=1
    wb.save('Horraires_GTFS.xlsx')

try:
    print('Création de la base de données...')
    connection = sqlite3.connect('BD_GTFS.db')
    print('Connection à la base de donnée établie.')
    print('-'*50)

    print('Récupération des fichiers et créations des tables...')
    necessaryFiles = ['routes.txt', 'stop_times.txt', 'stops.txt','trips.txt', 'calendar.txt']
    
    
    #Creation Tables of data_base
    for csvfile in necessaryFiles :
        print('   -Table {}...'.format( csvfile[:-4]))
        createDataBaseTableFromCsv(csvfile)
        print('        -Table {} crée'.format(csvfile[:-4]))
        
    print('-'*50)    
    print('Récupération des routes...')
    cursor = connection.execute("SELECT route_id FROM routes") #Je récupère la liste de toutes les routes
    listRoutes = [row[0] for row in cursor]
    print('Routes trouvées : ',listRoutes)
    
    print('-'*50)
    print('Création des dictionnaires...')
    print('    -Etape 1/4...')
    dictTrip0 = extractTrip(listRoutes, 0) #j'extrait tout les voyages par ligne dans le sens 0
    print('    -Etape 2/4...')
    dictTrip1 = extractTrip(listRoutes, 1) #j'extrait tout les voyages par ligne dans le sens 1
    print('    -Etape 3/4...')
    result = suppdoubleAndExtractStops(dictTrip0, listRoutes)
    dictTripStops0, dicDays0, dicPeriodRoute0 = result[0], result[1], result[2]
    print('    -Etape 4/4...')
    result = suppdoubleAndExtractStops(dictTrip1, listRoutes)
    dictTripStops1, dicDays1, dicPeriodRoute1 = result[0], result[1], result[2]
    print('-'*50)
    print('Création des fichiers .csv...')
    createCSV(dictTripStops0, dictTripStops1, dicDays0, dicDays1, dicPeriodRoute0, dicPeriodRoute1)
    print('-'*50)
    print('Création du fichier .xlsx...')
    createXLS(listRoutes, dictTripStops0, dictTripStops1, dicDays0, dicDays1, dicPeriodRoute0, dicPeriodRoute1)
    print('-'*50)
    print("L'oppération c'est déroulée correctement")
    
except sqlite3.Error as error:
    print('Une erreur est survenue -', error)
    
except FileNotFoundError:
    print('Erreur ! Fichier "{}" manquant.'.format(csvfile))

except KeyboardInterrupt:
    print("Arret de l'oppération")

finally:
    if connection :
        connection.close()
        print('La connection à la base de donnée à été fermée.')    
    os.system("pause")
